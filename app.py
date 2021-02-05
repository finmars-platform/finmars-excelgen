#! /usr/bin/env python

import sys
import os
import time

import json
import yaml


from werkzeug.wsgi import wrap_file
from werkzeug.wrappers import Request, Response
from executor import execute

from werkzeug.datastructures import Headers


from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.writer.excel import save_virtual_workbook


import logging

logger = logging.getLogger()

gunicorn_error_logger = logging.getLogger('gunicorn.error')
logger.handlers.extend(gunicorn_error_logger.handlers)
logger.setLevel(logging.DEBUG)



@Request.application
def application(request):

    if request.method == 'GET':
        return Response("Excel generator is online")

    if request.method != 'POST':
        return

    st = time.perf_counter()

    # logger.info("Request data %s" % request.data)

    # print('request.data %s' %request.data)

    # requestData = json.loads(request.data)

    requestData = yaml.safe_load(request.data)

    contentSettings = requestData['contentSettings']
    content = requestData['content']

    print("Creating empty workbook")
    wb = Workbook()
    ws = wb.active

    columns = contentSettings["columns"]

    column_names = []

    colorFill = PatternFill(start_color='F05A22',
                            end_color='F05A22',
                            fill_type='solid')

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ft = Font(color="FFFFFF", name='Arial', size=14)


    for column in columns:
        column_names.append(column['name'])

    print("Adding columns")
    ws.append(column_names)

    for col in ws.iter_cols(min_row=1, max_col=len(columns), max_row=1):
        for cell in col:
            cell.fill = colorFill
            cell.font = ft
            cell.border = thin_border

    print("Adding data")

    index = 0

    for row in content:

        row_data = []
        skip = False

        if '___type' in row and row['___type'] == 'subtotal':
            skip = True

        for column in columns:

            try:

                row_data.append(row[column["key"]])

            except Exception as e:
                row_data.append(None)

        if not skip:
            ws.append(row_data)

        index = index + 1


    print("Saving to excel file")

    # wb.save(filename='report.xlsx')

    headers = Headers()
    headers.add('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    headers.add('Content-Disposition', 'attachment', filename='report.xslx')

    # file = open('report.xslx')
    #
    # os.remove('report.xslx')


    response = Response(save_virtual_workbook(wb),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers
                        )

    print('generating excel done: %s', "{:3.3f}".format(time.perf_counter() - st))

    # file.close()

    return response


if __name__ == '__main__':
    from werkzeug.serving import run_simple

    run_simple(
        '127.0.0.1', 5000, application, use_debugger=True, use_reloader=True
    )
